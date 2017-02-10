import os
import re
import pyodbc
import openpyxl

# SELECT PARA OS CODIGOS DE PRODUTOS
_select_produto = "select cod_prod as codigo, desc_prod " \
                  "from produtos  " \
                  "WHERE cod_cat <> 29 and cod_cat<> 57"

# CONEXAO COM O BANCO DE DADOS
conn = pyodbc.connect(r'DRIVER={SQL Server};SERVER=server2008\;DATABASE=techcd;Trusted_Connection=True;')
cursor = conn.cursor()

# RETRIVE DOS DADOS
cursor.execute(_select_produto)
rows = cursor.fetchall()

# LISTA OS ARQUIVOS NO DIRETORIO
list_all_files = os.listdir(r'\\file_server\fotos')
list_thumbnail = list()
list_normal = list()

# nomes das planilhas utilizado
prod_sem_foto ='Prod sem foto'
prod_sem_thumnail = 'Prod sem thumbnail'
prod_sem_normal_foto = 'Prod sem foto normal'

# listas ja com arquivos thumbnail e padrao separados
thumbnail = [file[:-5] for file in list_all_files if re.search('(p{1})(.gif)', file)]
file_normal = [file[:-4] for file in list_all_files if re.search('([0-9])(.gif)', file)]


book = openpyxl.Workbook()
columns = ('CODIGO', 'DESCRICAO')

# produto sem nenhuma foto
sheet = book.active
sheet.title = prod_sem_foto
sheet.append(columns)
for row in rows:
    if str(row.codigo) not in thumbnail and str(row.codigo) not in file_normal:
        data = list(row)
        sheet.append(data)

# produto sem thumbnail
book.create_sheet(title=prod_sem_thumnail)
sheet = book.get_sheet_by_name(prod_sem_thumnail)
sheet.append(columns)
for row in rows:
    if str(row.codigo) not in thumbnail:
        data = list(row)
        sheet.append(data)

# produto sem foto padrao
book.create_sheet(title=prod_sem_normal_foto)
sheet = book.get_sheet_by_name(prod_sem_normal_foto)
sheet.append(columns)
for row in rows:
    if str(row.codigo) not in file_normal:
        data = list(row)
        sheet.append(data)

book.save(r'c:\output\output.xlsx')
