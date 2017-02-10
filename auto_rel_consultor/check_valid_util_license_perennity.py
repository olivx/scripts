import openpyxl
import warnings
import smtplib
import os

from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


''' para não executar warnigs irrtantes '''
warnings.filterwarnings('ignore')

''' path onde fica o xlsx '''
# os.system('sudo mount -t cifs -o username=thiago.oliveira,password=logan //file_server/sistemas /mnt/sistemas')
# xls_file = u'/mnt/sistemas/suporte/valid_util_perennity.xlsx'
xls_file = u'/home/olvx/Área de Trabalho/valid_util_perennity.xlsx'

xls = openpyxl.load_workbook(xls_file)


''' aqui vou para os dicionarios com os valores '''
list_rows  = []
list_danger_valid =[ 0, 1 ,2, 3, 4, 5]
list_rows_i_need_send = []

''' chaves para o seu dicionario '''
_INDEX =  'INDEX'
_CLIENTE =  'CLIENTE'
_TECNICO = 'TECNICO'
_VALID =  'VALID_UTIL'
_SERIAL_NUMBER = 'SERIAL_NUMBER'
_LICEMSE = 'LICENSE'
_MAC_ADRESS = 'MAC_ADRESS'
_TYPE = 'TYPE'
_SENT = 'SENT'


''' colunas chaves para os busca de dados '''
column_clientes = 1
column_tecnico = 2
column_valid_util = 3
column_serial_number = 4
column_mac_adress = 5
column_type = 6
column_license = 7
column_sent = 8



''' referencia a plhanilha para leitura '''
sheet = xls.get_sheet_by_name('dados')

''' loop  na planilha '''
# for position in range(2,sheet.max_row):
for position in range(2,sheet.get_highest_row()):
    ''' inicializa o dicionario com dados da planilha'''
    dict_values = {}

    ''' input dos dados das linceças na planilha '''
    dict_values[_INDEX] = position
    dict_values[_CLIENTE] =  sheet.cell(row = position, column = column_clientes).value
    dict_values[_TECNICO] =  sheet.cell(row = position, column = column_tecnico).value
    dict_values[_VALID] =  sheet.cell(row = position, column = column_valid_util).value
    dict_values[_SERIAL_NUMBER] =  sheet.cell(row = position, column = column_serial_number).value
    dict_values[_MAC_ADRESS] =  sheet.cell(row = position, column = column_mac_adress).value
    dict_values[_TYPE] =  sheet.cell(row = position, column = column_type).value
    dict_values[_LICEMSE] =  sheet.cell(row = position, column = column_license).value
    dict_values[_SENT] =  sheet.cell(row = position, column = column_sent).value

    ''' storage na  lista '''
    list_rows.append(dict_values)


'''  todo

verificar a data está proxima da expirar , caso esteja para expirar em 5 dias

enviar email enformando para souporte@techcd.com.br , depois de expirada marcar a

coluna como expirada e tirar a colunada sent para saber que o e-mail foi enviado.


'''


'''exemplo  varre a lista '''
for _list in list_rows:

    ''' dicionario atual '''
    dict_values = _list

    ''' aqui vai acontecer a magica !'''
    # for key, values in dict_values.items():
    now = datetime.now()

    ''' se eu  '''
    diff_off = dict_values[_VALID] - now
    if  diff_off.days in list_danger_valid and dict_values[_SENT] is not None:

        ''' dicionario com dados preste a vencer que será enviado por e-mail para o suporte '''
        list_rows_i_need_send.append(dict_values)





# ''' envia o email com as informações corretas '''
# from_addr = 'thiago@techcd.com.br'
# to_addr = 'thiago@techcd.com.br'
# recipents =  ['thiago@techcd.com.br', 'oliveiravicente.net@gmail.com']
#
# message = MIMEMultipart()
# message['From'] = from_addr
# message['To'] = to_addr
# message['Cc'] = ','.join(recipents)
# message['Subject'] = 'Controle de Inslação de Linsença Perennity'
#
# body  = '''
#
# Essa é uma messagem de teste de envio
#
# '''
#
#
# ''' menssagem a ser enviada '''
# message.attach(MIMEText(body, 'plain'))
#
# smtp  = smtplib.SMTP('smtp.techcd.com.br', 587)
# smtp.starttls()
# smtp.login('thiago@techcd.com.br', 'mmnhbn')
# text =  message.as_string()
# smtp.sendmail(from_addr, recipents,  text)
# smtp.quit()
